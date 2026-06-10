from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from .influx_pool import InfluxPool, InfluxQueryError
from .alert_monitor import severity_from_body
from .dashboard import (
    scoped_buckets,
    bucket_key,
    range_clause,
    strip_syslog_header,
    bucket_severity,
    _MEASUREMENT_FILTER,
    _COUNT_FIELD_FILTER,
)

log = logging.getLogger("unifiedops.reports")

def _flux_report(bucket: str, range_key: str, limit: int = 50000) -> str:
    """Pull raw messages up to a large limit for export."""
    return (
        f'from(bucket: "{bucket}")\n'
        f'  |> range({range_clause(range_key)})\n'
        f'  |> filter(fn: (r) => {_MEASUREMENT_FILTER})\n'
        f'  |> filter(fn: (r) => {_COUNT_FIELD_FILTER})\n'
        f'  |> sort(columns: ["_time"], desc: true)\n'
        f'  |> limit(n: {limit})\n'
    )

class ReportService:
    def __init__(self, pool: InfluxPool) -> None:
        self._pool = pool

    async def get_multi_format_report(
        self,
        range_key: str,
        sites: Optional[List[str]] = None,
        vendors: Optional[List[str]] = None,
        limit: int = 50000,
        fmt: str = "csv",
    ) -> tuple[bytes, str, str]:
        """Fetch alerts across requested sites/vendors and return as formatted bytes, along with media_type and extension."""
        import io
        import csv
        import zipfile
        from collections import defaultdict
        import openpyxl
        from fpdf import FPDF

        buckets = scoped_buckets(sites, vendors)
        per_bucket = await asyncio.gather(*[
            self._safe_query(bucket_key(b), _flux_report(b["bucket"], range_key, limit))
            for b in buckets
        ])

        out: List[Dict[str, Any]] = []
        for cfg, rows in zip(buckets, per_bucket):
            for r in rows:
                ts = r.get("_time")
                if not ts:
                    continue
                try:
                    dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                except Exception:
                    continue
                
                local_time = dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
                
                storage = (
                    r.get("array_name")
                    or r.get("switch_name")
                    or r.get("hostname")
                    or r.get("source_ip", "-")
                )
                if storage in ("", "unknown", "-"):
                    storage = r.get("source_ip", "-")
                
                raw_body = str(r.get("_value") or "")
                event_text = strip_syslog_header(raw_body) or raw_body
                
                raw_sev = r.get("severity")
                severity = bucket_severity(raw_sev) if raw_sev else "informational"
                if not raw_sev or severity == "informational":
                    body_sev = severity_from_body(raw_body)
                    if body_sev is not None:
                        severity = body_sev
                
                out.append({
                    "Timestamp": ts,
                    "Local Time": local_time,
                    "Severity": severity.capitalize(),
                    "Storage/Switch": storage,
                    "Source IP": r.get("source_ip") or "-",
                    "Event Details": event_text,
                    "Category": (r.get("trap_category") or "other").capitalize(),
                    "Location": cfg["site"],
                    "Vendor": cfg["vendor"].capitalize(),
                })
        
        # Sort by actual timestamp descending
        out.sort(key=lambda a: a["Timestamp"], reverse=True)

        fieldnames = [
            "Local Time", "Location", "Vendor", "Storage/Switch",
            "Severity", "Category", "Source IP", "Event Details"
        ]

        if not out:
            if fmt == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "No Data"
                ws.append(["No data found for the selected criteria."])
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
            elif fmt == "pdf":
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr("No_Data.txt", b"No data found for the selected criteria.")
                return buf.getvalue(), "application/zip", "zip"
            else:
                return b"No data found for the selected criteria.\n", "text/csv", "csv"

        # Group data by Vendor_Site
        groups = defaultdict(list)
        for row in out:
            tab_name = f"{row['Vendor']}_{row['Location']}"
            groups[tab_name].append(row)

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            # Remove default active sheet if we are going to create our own
            if len(wb.sheetnames) > 0:
                del wb[wb.sheetnames[0]]

            for tab_name, rows in groups.items():
                ws = wb.create_sheet(title=tab_name[:31])
                ws.append(fieldnames)
                for r in rows:
                    ws.append([str(r[f]) for f in fieldnames])

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

        elif fmt == "pdf":
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for tab_name, rows in groups.items():
                    pdf = FPDF(orientation="landscape")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.add_page()
                    
                    pdf.set_font("helvetica", "B", 12)
                    pdf.cell(0, 10, f"UnifiedOps Alerts: {tab_name}", new_x="LMARGIN", new_y="NEXT", align="C")
                    pdf.ln(5)
                    
                    # Define headers to display in PDF
                    pdf_headers = ["Local Time", "Severity", "Storage/Switch", "Source IP", "Category", "Event Details"]
                    
                    with pdf.table(col_widths=(30, 20, 40, 25, 25, 137), text_align="LEFT") as table:
                        header_row = table.row()
                        pdf.set_font("helvetica", "B", 9)
                        for header in pdf_headers:
                            header_row.cell(header)
                        
                        pdf.set_font("helvetica", "", 8)
                        for r in rows:
                            data_row = table.row()
                            data_row.cell(str(r["Local Time"]))
                            data_row.cell(str(r["Severity"]))
                            data_row.cell(str(r["Storage/Switch"]))
                            data_row.cell(str(r["Source IP"]))
                            data_row.cell(str(r["Category"]))
                            data_row.cell(str(r["Event Details"]))
                        
                    pdf_bytes = pdf.output()
                    zf.writestr(f"{tab_name}.pdf", pdf_bytes)

            return buf.getvalue(), "application/zip", "zip"

        else:
            # Default to CSV
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for row in out:
                writer.writerow(row)
            return output.getvalue().encode('utf-8'), "text/csv", "csv"

    async def _safe_query(self, key: str, flux: str) -> List[Dict[str, Any]]:
        try:
            return await self._pool.query(key, flux)
        except InfluxQueryError as exc:
            log.warning("reports %s query failed: %s", key, exc.reason)
            return []
        except Exception as exc:
            log.warning("reports %s query crash: %s", key, exc)
            return []
